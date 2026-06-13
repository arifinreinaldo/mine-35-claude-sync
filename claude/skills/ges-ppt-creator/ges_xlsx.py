#!/usr/bin/env python3
"""Question-based Excel (.xlsx) intake for ges-ppt-creator.

The workbook has three columns: Question (human prompt), Answer (user fills),
and Key (hidden, used for mapping). One row per atomic answer; list items get
numbered keys (challenge.1.item.2, scope.3.module, commercial.2.unit_price, ...).
Read by the Key/Answer header names, so column order/reordering is tolerated.

Usage:
    python3 ges_xlsx.py blank out.xlsx
    python3 ges_xlsx.py fromjson intake.json out.xlsx [--blank-prices]
    python3 ges_xlsx.py tojson filled.xlsx out.json
"""
import sys, json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def _num(s):
    s = str(s if s is not None else "").replace(",", "").replace(".", "").replace(" ", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


# ---------- intake -> ordered [(key, question, value)] ----------
def _flatten(intake, blank_prices=False):
    rows, c = [], intake.get("company", {})
    add = lambda k, q, v="": rows.append((k, q, "" if v is None else v))
    add("company.name", "Client company name", c.get("name", ""))
    add("company.industry", "Industry", c.get("industry", ""))
    add("company.size", "Company size (e.g. '50 employees, 2 branches')", c.get("size", ""))
    add("company.current_system", "Current system(s) used today", c.get("current_system", ""))
    add("company.date", "Proposal date (e.g. 'July 2026')", c.get("date", ""))
    u = intake.get("understanding", {})
    add("understanding.profile", "Company profile (one short paragraph)", u.get("profile", ""))
    add("understanding.current_system", "Current-state note (Customer Profile column)", u.get("current_system", ""))
    chs = u.get("challenges", [])
    for i in range(max(3, len(chs))):
        ch = chs[i] if i < len(chs) else {"category": "", "items": []}
        add("challenge.%d.category" % (i + 1), "Challenge group %d — title" % (i + 1), ch.get("category", ""))
        items = ch.get("items", [])
        for j in range(max(3, len(items))):
            add("challenge.%d.item.%d" % (i + 1, j + 1), "Challenge group %d — point %d" % (i + 1, j + 1),
                items[j] if j < len(items) else "")
    stages = intake.get("process", {}).get("stages", [])
    for i in range(max(4, len(stages))):
        st = stages[i] if i < len(stages) else {"title": "", "steps": []}
        add("process.%d.title" % (i + 1), "Process stage %d — title" % (i + 1), st.get("title", ""))
        steps = st.get("steps", [])
        for j in range(max(4, len(steps))):
            add("process.%d.step.%d" % (i + 1, j + 1), "Process stage %d — step %d" % (i + 1, j + 1),
                steps[j] if j < len(steps) else "")
    add("process.loop_note", "Process — closed-loop note", intake.get("process", {}).get("loop_note", ""))
    add("process.reporting_note", "Process — reporting / dashboard note", intake.get("process", {}).get("reporting_note", ""))
    bens = intake.get("benefits", [])
    for i in range(max(8, len(bens))):
        add("benefit.%d" % (i + 1), "Business benefit %d" % (i + 1), bens[i] if i < len(bens) else "")
    m = intake.get("methodology", {})
    add("methodology.start_month", "Project start month (e.g. 'August')", m.get("start_month", ""))
    add("methodology.months", "Project duration (months)", m.get("months", ""))
    mst = m.get("stages", [])
    for i in range(max(6, len(mst))):
        add("methodology.stage.%d" % (i + 1), "Methodology stage %d" % (i + 1), mst[i] if i < len(mst) else "")
    scope = intake.get("scope", [])
    for i in range(max(6, len(scope))):
        s = scope[i] if i < len(scope) else {}
        n = i + 1
        add("scope.%d.module" % n, "Scope %d — module" % n, s.get("module", ""))
        add("scope.%d.solution" % n, "Scope %d — 1C:Drive solution" % n, s.get("solution", ""))
        add("scope.%d.type" % n, "Scope %d — type (Standard / Enhancement / Standard, Enhancement)" % n, s.get("type", ""))
        add("scope.%d.scope_function" % n, "Scope %d — scope function (what it covers)" % n, s.get("scope_function", ""))
    com = intake.get("commercials", {})
    add("commercial.currency", "Commercials — currency (e.g. IDR)", com.get("currency", ""))
    add("commercial.terms", "Commercials — payment terms", com.get("terms", ""))
    items = com.get("line_items", [])
    for i in range(max(4, len(items))):
        li = items[i] if i < len(items) else {}
        n = i + 1
        add("commercial.%d.item" % n, "Commercial line %d — item" % n, li.get("item", ""))
        add("commercial.%d.qty" % n, "Commercial line %d — quantity" % n, li.get("qty", "") if li else "")
        price = "" if (blank_prices or not li.get("unit_price")) else li["unit_price"]
        add("commercial.%d.unit_price" % n, "Commercial line %d — unit price (numbers only)" % n, price)
    asm = intake.get("assumptions", [])
    for i in range(max(6, len(asm))):
        add("assumption.%d" % (i + 1), "Assumption %d" % (i + 1), asm[i] if i < len(asm) else "")
    return rows


# ---------- [(key, answer)] -> intake ----------
def _unflatten(pairs):
    it = {"company": {}, "understanding": {"challenges": []}, "process": {"stages": []},
          "benefits": [], "methodology": {"stages": []}, "scope": [],
          "commercials": {"line_items": []}, "assumptions": []}
    ch, st, sc, ci = {}, {}, {}, {}
    ben, asm, mst = {}, {}, {}
    for key, val in pairs:
        val = "" if val is None else str(val).strip()
        p = key.split(".")
        sec = p[0]
        if sec == "company":
            it["company"][p[1]] = val
        elif sec == "understanding":
            it["understanding"][p[1]] = val
        elif sec == "challenge":
            c = ch.setdefault(int(p[1]), {"category": "", "items": {}})
            if p[2] == "category":
                c["category"] = val
            elif p[2] == "item" and val:
                c["items"][int(p[3])] = val
        elif sec == "process":
            if p[1] in ("loop_note", "reporting_note"):
                it["process"][p[1]] = val
            else:
                s = st.setdefault(int(p[1]), {"title": "", "steps": {}})
                if p[2] == "title":
                    s["title"] = val
                elif p[2] == "step" and val:
                    s["steps"][int(p[3])] = val
        elif sec == "benefit":
            if val:
                ben[int(p[1])] = val
        elif sec == "methodology":
            if p[1] == "stage":
                if val:
                    mst[int(p[2])] = val
            else:
                it["methodology"][p[1]] = _num(val) if p[1] == "months" else val
        elif sec == "scope":
            sc.setdefault(int(p[1]), {})[p[2]] = val
        elif sec == "commercial":
            if p[1] in ("currency", "terms"):
                it["commercials"][p[1]] = val
            else:
                ci.setdefault(int(p[1]), {})[p[2]] = val
        elif sec == "assumption":
            if val:
                asm[int(p[1])] = val
    # assemble ordered lists
    for i in sorted(ch):
        c = ch[i]
        if c["category"] or c["items"]:
            it["understanding"]["challenges"].append(
                {"category": c["category"], "items": [c["items"][j] for j in sorted(c["items"])]})
    for i in sorted(st):
        s = st[i]
        if s["title"] or s["steps"]:
            it["process"]["stages"].append(
                {"title": s["title"], "steps": [s["steps"][j] for j in sorted(s["steps"])]})
    it["benefits"] = [ben[i] for i in sorted(ben)]
    it["methodology"]["stages"] = [mst[i] for i in sorted(mst)]
    for i in sorted(sc):
        s = sc[i]
        if s.get("module"):
            it["scope"].append({"module": s.get("module", ""), "solution": s.get("solution", ""),
                                 "type": s.get("type", ""), "scope_function": s.get("scope_function", "")})
    for i in sorted(ci):
        li = ci[i]
        if li.get("item"):
            it["commercials"]["line_items"].append(
                {"item": li.get("item", ""), "qty": _num(li.get("qty")) or 1, "unit_price": _num(li.get("unit_price"))})
    it["assumptions"] = [asm[i] for i in sorted(asm)]
    return it


# ---------- workbook I/O ----------
def write_xlsx(intake, path, blank_prices=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proposal intake"
    ws.append(["Question", "Answer", "Key"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="21465E")
    for k, q, v in _flatten(intake, blank_prices):
        ws.append([q, v, k])
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].hidden = True
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].fill = PatternFill("solid", fgColor="FFF8E1")  # highlight the Answer column
    wb.save(path)
    return path


def xlsx_to_intake(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    ki, ai = header.index("Key"), header.index("Answer")
    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[ki] if ki < len(row) else None
        ans = row[ai] if ai < len(row) else None
        if key:
            pairs.append((str(key).strip(), ans))
    return _unflatten(pairs)


_BLANK = {
    "company": {}, "understanding": {"challenges": []}, "process": {"stages": []},
    "benefits": [], "methodology": {"start_month": "", "months": 6,
                                    "stages": ["Impact Analysis", "Functional Design", "Customization",
                                               "Implementation", "Training", "Go-live"]},
    "scope": [], "commercials": {"currency": "IDR", "terms": "",
                                 "line_items": [{"item": n, "qty": q, "unit_price": 0} for n, q in
                                                (("1C:Drive named-user licenses", 10), ("Implementation services", 1),
                                                 ("Custom enhancements", 1), ("Annual support & maintenance (Year 1)", 1))]},
    "assumptions": [],
}


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "blank"
    if cmd == "blank":
        print("wrote", write_xlsx(_BLANK, sys.argv[2], blank_prices=True))
    elif cmd == "fromjson":
        print("wrote", write_xlsx(json.load(open(sys.argv[2])), sys.argv[3], blank_prices="--blank-prices" in sys.argv))
    elif cmd == "tojson":
        json.dump(xlsx_to_intake(sys.argv[2]), open(sys.argv[3], "w"), indent=2, ensure_ascii=False)
        print("wrote", sys.argv[3])
